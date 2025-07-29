#!/usr/bin/env python3
"""
Config File Matrix Tool

Processes configuration files organized in hostname-based folders and creates
a matrix-style Excel report showing configuration values across different hostnames.
Includes hostname intelligence to identify when differences are only due to hostname numbering.
"""

import argparse
import re
import sys
import pandas as pd
from typing import Dict, List, Optional, Set, Tuple
from pathlib import Path
import glob
import os
import xml.etree.ElementTree as ET


def parse_config_file(file_path: str) -> Dict[str, str]:
    """
    Parse a configuration file and extract key-value pairs.
    Supports .conf, .rc (key=value format) and .xml files.
    
    Args:
        file_path: Path to the configuration file
        
    Returns:
        Dictionary of key-value pairs
    """
    config_dict = {}
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if file_ext == '.xml':
            return parse_xml_file(file_path)
        else:
            # Handle .conf and .rc files (key=value format)
            return parse_key_value_file(file_path)
            
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return {}
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        return {}


def parse_key_value_file(file_path: str) -> Dict[str, str]:
    """
    Parse a key=value format configuration file (.conf, .rc).
    
    Args:
        file_path: Path to the configuration file
        
    Returns:
        Dictionary of key-value pairs
    """
    config_dict = {}
    
    with open(file_path, 'r', encoding='utf-8') as file:
        for line_num, line in enumerate(file, 1):
            line = line.strip()
            
            # Skip empty lines, lines starting with # or space
            if not line or line.startswith('#') or line.startswith(' '):
                continue
            
            # Split on first '=' character
            if '=' in line:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                config_dict[key] = value
                
    return config_dict


def parse_xml_file(file_path: str) -> Dict[str, str]:
    """
    Parse an XML configuration file and extract key-value pairs.
    Flattens XML structure using dot notation for nested elements.
    
    Args:
        file_path: Path to the XML configuration file
        
    Returns:
        Dictionary of key-value pairs with dot notation for nested elements
    """
    config_dict = {}
    
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        def extract_elements(element, prefix=""):
            # Handle element text content
            if element.text and element.text.strip():
                key = f"{prefix}{element.tag}" if prefix else element.tag
                config_dict[key] = element.text.strip()
            
            # Handle element attributes
            if element.attrib:
                for attr_name, attr_value in element.attrib.items():
                    key = f"{prefix}{element.tag}.@{attr_name}" if prefix else f"{element.tag}.@{attr_name}"
                    config_dict[key] = attr_value
            
            # Handle child elements
            for child in element:
                child_prefix = f"{prefix}{element.tag}." if prefix else f"{element.tag}."
                extract_elements(child, child_prefix)
        
        extract_elements(root)
        
    except ET.ParseError as e:
        print(f"Error parsing XML file '{file_path}': {e}")
        return {}
    
    return config_dict


def normalize_hostname(value: str) -> str:
    """
    Normalize hostname by replacing numeric suffixes with a placeholder.
    
    This helps identify when the only difference is hostname numbering.
    Examples:
        apesap-h-koc-1 -> apesap-h-koc-X
        apesap-h-top-2 -> apesap-h-top-X
        server-web-01 -> server-web-X
        db-cluster-03 -> db-cluster-X
    """
    # Pattern to match hostname with numeric suffix (more flexible patterns)
    patterns = [
        r'([a-zA-Z]+-[a-zA-Z]+-[a-zA-Z]+-)\d+$',  # Original pattern: word-word-word-number
        r'([a-zA-Z]+-[a-zA-Z]+-)\d+$',            # word-word-number
        r'([a-zA-Z]+)\d+$',                       # word+number
        r'([a-zA-Z]+-[a-zA-Z]+-[a-zA-Z]+-[a-zA-Z]+-)\d+$'  # word-word-word-word-number
    ]
    
    for pattern in patterns:
        match = re.match(pattern, value)
        if match:
            return match.group(1) + 'X'
    
    return value


def is_hostname_only_difference(value1: str, value2: str) -> bool:
    """
    Check if the only difference between two values is the hostname numbering.
    
    Args:
        value1: First value to compare
        value2: Second value to compare
        
    Returns:
        True if only difference is hostname numbering, False otherwise
    """
    if value1 == value2:
        return False
        
    normalized1 = normalize_hostname(value1)
    normalized2 = normalize_hostname(value2)
    
    return normalized1 == normalized2 and value1 != value2


def analyze_hostname_differences(hostname_configs: Dict[str, Dict[str, str]]) -> Dict[str, Dict[str, List[str]]]:
    """
    Analyze configurations across hostnames to identify hostname-only differences.
    
    Args:
        hostname_configs: Dictionary mapping hostname to config dict
        
    Returns:
        Dictionary with analysis results: {
            'hostname_only_differences': {config_key: [list of hostnames with differences]},
            'significant_differences': {config_key: [list of hostnames with differences]}
        }
    """
    result = {
        'hostname_only_differences': {},
        'significant_differences': {}
    }
    
    if len(hostname_configs) < 2:
        return result
    
    # Get all config keys
    all_keys = set()
    for config in hostname_configs.values():
        all_keys.update(config.keys())
    
    hostnames = list(hostname_configs.keys())
    
    for key in all_keys:
        # Get all values for this key across hostnames
        values = {}
        for hostname in hostnames:
            if hostname in hostname_configs and key in hostname_configs[hostname]:
                values[hostname] = hostname_configs[hostname][key]
        
        if len(values) < 2:
            continue
            
        # Check if all differences are hostname-only
        hostname_only = True
        significant_diff = False
        
        hostnames_with_values = list(values.keys())
        for i in range(len(hostnames_with_values)):
            for j in range(i + 1, len(hostnames_with_values)):
                host1, host2 = hostnames_with_values[i], hostnames_with_values[j]
                val1, val2 = values[host1], values[host2]
                
                if val1 != val2:
                    significant_diff = True
                    if not is_hostname_only_difference(val1, val2):
                        hostname_only = False
        
        if significant_diff:
            if hostname_only:
                result['hostname_only_differences'][key] = hostnames_with_values
            else:
                result['significant_differences'][key] = hostnames_with_values
    
    return result


def find_config_files(base_path: str) -> Dict[str, Dict[str, Dict[str, str]]]:
    """
    Find all configuration files in the hostname-based folder structure.
    Supports .conf, .rc, and .xml files in both direct hostname folders
    and subdirectories like data_explorer/, rc/, site/.
    
    Args:
        base_path: Base path containing server type folders
        
    Returns:
        Dictionary mapping server_type to hostname to dict of {relative_path: filepath}
        Format: {server_type: {hostname: {relative_path: filepath}}}
    """
    server_hostname_files = {}
    
    # Supported file extensions
    supported_extensions = ['*.conf', '*.rc', '*.xml']
    
    # Look for folders in the pattern SERVER_TYPE/hostname
    pattern = os.path.join(base_path, "*", "*")
    folders = glob.glob(pattern)
    
    for folder_path in folders:
        if os.path.isdir(folder_path):
            # Extract hostname from path (last part)
            hostname = os.path.basename(folder_path)
            server_type = os.path.basename(os.path.dirname(folder_path))
            
            # Find all supported config files in this hostname folder and subdirectories
            config_files = []
            
            # Search in the hostname folder directly
            for ext in supported_extensions:
                config_files.extend(glob.glob(os.path.join(folder_path, ext)))
            
            # Search in common subdirectories (recursive search up to 3 levels deep)
            for ext in supported_extensions:
                config_files.extend(glob.glob(os.path.join(folder_path, "*", ext)))
                config_files.extend(glob.glob(os.path.join(folder_path, "*", "*", ext)))
                config_files.extend(glob.glob(os.path.join(folder_path, "*", "*", "*", ext)))
            
            if config_files:
                if server_type not in server_hostname_files:
                    server_hostname_files[server_type] = {}
                if hostname not in server_hostname_files[server_type]:
                    server_hostname_files[server_type][hostname] = {}
                
                for config_file in config_files:
                    # Create relative path from hostname folder to maintain hierarchy info
                    rel_path = os.path.relpath(config_file, folder_path)
                    # Replace path separators with forward slash for consistency
                    rel_path = rel_path.replace(os.sep, '/')
                    server_hostname_files[server_type][hostname][rel_path] = config_file
    
    return server_hostname_files


def create_matrix_data(hostname_files: Dict[str, Dict[str, str]], enable_hostname_intelligence: bool = True) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Create matrix data with filenames as rows and hostnames as columns.
    
    Args:
        hostname_files: Dictionary mapping hostname to {filename: filepath}
        enable_hostname_intelligence: Whether to enable hostname intelligence analysis
        
    Returns:
        Tuple of (DataFrame with matrix structure, Dictionary with hostname analysis)
    """
    # Get all unique filenames across all hostnames
    all_filenames = set()
    all_hostnames = list(hostname_files.keys())
    
    for hostname_data in hostname_files.values():
        all_filenames.update(hostname_data.keys())
    
    all_filenames = sorted(list(all_filenames))
    all_hostnames = sorted(all_hostnames)
    
    # Create matrix structure
    matrix_data = []
    hostname_analysis = {}
    
    for filename in all_filenames:
        # Get all config keys from this filename across all hostnames
        all_keys = set()
        hostname_configs = {}
        
        for hostname in all_hostnames:
            if hostname in hostname_files and filename in hostname_files[hostname]:
                config = parse_config_file(hostname_files[hostname][filename])
                hostname_configs[hostname] = config
                all_keys.update(config.keys())
        
        # Analyze hostname differences for this file (if enabled)
        if enable_hostname_intelligence:
            file_analysis = analyze_hostname_differences(hostname_configs)
        else:
            file_analysis = {'hostname_only_differences': {}, 'significant_differences': {}}
        
        # Create rows for each config key in this file
        for key in sorted(all_keys):
            row = {'File': filename, 'Config_Key': key}
            
            # Determine if this config key has hostname-only differences (if enabled)
            if enable_hostname_intelligence:
                difference_type = ''
                if key in file_analysis['hostname_only_differences']:
                    difference_type = 'Hostname-Only'
                elif key in file_analysis['significant_differences']:
                    difference_type = 'Significant'
                else:
                    difference_type = 'Same'
                
                row['Difference_Type'] = difference_type
            
            for hostname in all_hostnames:
                if hostname in hostname_configs and key in hostname_configs[hostname]:
                    row[hostname] = hostname_configs[hostname][key]
                else:
                    row[hostname] = ''
            
            matrix_data.append(row)
            
            # Store analysis for summary (if enabled)
            if enable_hostname_intelligence:
                full_key = f"{filename}:{key}"
                hostname_analysis[full_key] = difference_type
    
    return pd.DataFrame(matrix_data), hostname_analysis


def write_matrix_to_excel(matrix_df: pd.DataFrame, output_path: str, hostname_files: Dict[str, Dict[str, str]], hostname_analysis: Dict[str, str], server_type: str = ""):
    """
    Write matrix data to an Excel file with proper formatting and hostname intelligence.
    
    Args:
        matrix_df: DataFrame containing the matrix data
        output_path: Path for output Excel file
        hostname_files: Original hostname files mapping for summary
        hostname_analysis: Dictionary with hostname difference analysis
        server_type: Server type for this matrix (for display purposes)
    """
    try:
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write matrix to main sheet
            matrix_df.to_excel(writer, sheet_name='Config_Matrix', index=False)
            
            # Create summary sheet
            summary_data = []
            summary_data.append(['Metric', 'Value'])
            summary_data.append(['Total Hostnames', len(hostname_files)])
            summary_data.append(['Total Unique Files', len(matrix_df['File'].unique()) if not matrix_df.empty else 0])
            summary_data.append(['Total Config Keys', len(matrix_df) if not matrix_df.empty else 0])
            summary_data.append(['', ''])
            
            # Add hostname intelligence analysis
            if hostname_analysis:
                hostname_only_count = sum(1 for v in hostname_analysis.values() if v == 'Hostname-Only')
                significant_count = sum(1 for v in hostname_analysis.values() if v == 'Significant')
                same_count = sum(1 for v in hostname_analysis.values() if v == 'Same')
                
                summary_data.append(['Hostname Intelligence Analysis:', ''])
                summary_data.append(['  Same across hostnames', same_count])
                summary_data.append(['  Hostname-only differences', hostname_only_count])
                summary_data.append(['  Significant differences', significant_count])
                summary_data.append(['', ''])
            
            summary_data.append(['Hostnames Found:', ''])
            
            for hostname in sorted(hostname_files.keys()):
                summary_data.append([f'  {hostname}', f"{len(hostname_files[hostname])} files"])
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format Config_Matrix sheet
            if 'Config_Matrix' in workbook.sheetnames:
                worksheet = workbook['Config_Matrix']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add conditional formatting for hostname intelligence
                from openpyxl.styles import PatternFill
                hostname_only_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
                significant_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")   # Light red
                
                # Apply formatting based on Difference_Type column
                if 'Difference_Type' in matrix_df.columns:
                    diff_type_col = matrix_df.columns.get_loc('Difference_Type') + 1  # +1 for Excel 1-indexing
                    
                    for row_idx in range(2, len(matrix_df) + 2):  # Skip header row
                        cell_value = worksheet.cell(row=row_idx, column=diff_type_col).value
                        if cell_value == 'Hostname-Only':
                            # Highlight the entire row with light yellow
                            for col_idx in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = hostname_only_fill
                        elif cell_value == 'Significant':
                            # Highlight the entire row with light red
                            for col_idx in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = significant_fill
                
                # Freeze the first three columns (File, Config_Key, Difference_Type)
                worksheet.freeze_panes = worksheet['D2']
            
            # Format Summary sheet
            if 'Summary' in workbook.sheetnames:
                worksheet = workbook['Summary']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Excel matrix report generated: {output_path}")
        print(f"Matrix contains {len(matrix_df)} configuration entries across {len(hostname_files)} hostnames")
        
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        sys.exit(1)


def process_server_type_matrix(server_type: str, hostname_files: Dict[str, Dict[str, str]], base_output_path: str, enable_hostname_intelligence: bool = True):
    """
    Process matrix data for a specific server type and write to its own Excel file.
    
    Args:
        server_type: The server type (e.g., 'APP', 'DB', 'WEB')
        hostname_files: Dictionary mapping hostname to {filename: filepath} for this server type
        base_output_path: Base output path (without extension)
        enable_hostname_intelligence: Whether to enable hostname intelligence analysis
        
    Returns:
        Tuple of (matrix_df, hostname_analysis) for this server type
    """
    print(f"Processing {server_type} server type...")
    
    # Create matrix data for this server type
    matrix_df, hostname_analysis = create_matrix_data(hostname_files, enable_hostname_intelligence)
    
    if matrix_df.empty:
        print(f"  No configuration data found for {server_type} server type.")
        return matrix_df, hostname_analysis
    
    # Generate output filename for this server type
    base_name = os.path.splitext(base_output_path)[0]
    extension = os.path.splitext(base_output_path)[1] or '.xlsx'
    output_path = f"{base_name}_{server_type}_diff_matrix{extension}"
    
    # Write to Excel
    write_matrix_to_excel(matrix_df, output_path, hostname_files, hostname_analysis, server_type)
    
    print(f"  Generated: {output_path}")
    print(f"  Matrix contains {len(matrix_df)} configuration entries across {len(hostname_files)} hostnames")
    
    return matrix_df, hostname_analysis


def main():
    """Main function to run the config matrix tool."""
    parser = argparse.ArgumentParser(
        description="Create a matrix view of configuration files organized by hostname folders",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python config_diff_tool.py /path/to/configs
  python config_diff_tool.py /path/to/configs -o config_matrix.xlsx
  
Expected folder structure:
  base_path/
    ├── APP/
    │   ├── hostname1/
    │   │   ├── config1.conf
    │   │   ├── data_explorer/
    │   │   │   └── settings.xml
    │   │   └── rc/
    │   │       └── startup.rc
    │   └── hostname2/
    │       ├── config1.conf
    │       └── site/
    │           └── web.xml
    └── DB/
        ├── hostname1/
        │   ├── db.conf
        │   └── data_explorer/
        │       └── database.xml
        └── hostname2/
            └── db.conf

The tool will:
- Parse .conf, .rc (key=value format) and .xml files
- Search in hostname directories and subdirectories (data_explorer/, rc/, site/, etc.)
- For XML files: flatten structure using dot notation (e.g., server.database.host)
- For .conf/.rc files: parse key=value pairs, ignore lines starting with # or space
- Create a matrix with hostnames as columns and config keys as rows
- Show relative file path in the leftmost column
        """
    )
    
    parser.add_argument('base_path', help='Base path containing hostname folders')
    parser.add_argument('-o', '--output', default='config_matrix.xlsx',
                       help='Output Excel file path (default: config_matrix.xlsx)')
    parser.add_argument('--disable-hostname-intelligence', action='store_true',
                       help='Disable hostname intelligence analysis and highlighting')
    
    args = parser.parse_args()
    
    # Validate base path exists
    if not Path(args.base_path).exists():
        print(f"Error: Path '{args.base_path}' does not exist.")
        sys.exit(1)
    
    print(f"Scanning for configuration files in: {args.base_path}")
    print()
    
    # Find all config files organized by server type and hostname
    server_hostname_files = find_config_files(args.base_path)
    
    if not server_hostname_files:
        print("No configuration files found in the expected folder structure.")
        print("Expected structure: SERVER_TYPE/hostname/*.{conf,rc,xml}")
        print("Also searches subdirectories like data_explorer/, rc/, site/")
        sys.exit(1)
    
    print(f"Found configuration files for {len(server_hostname_files)} server types:")
    total_hostnames = 0
    total_files = 0
    for server_type, hostname_files in server_hostname_files.items():
        host_count = len(hostname_files)
        file_count = sum(len(files) for files in hostname_files.values())
        total_hostnames += host_count
        total_files += file_count
        print(f"  {server_type}: {host_count} hostnames, {file_count} total files")
    print()
    
    # Process each server type separately
    enable_hostname_intelligence = not args.disable_hostname_intelligence
    if enable_hostname_intelligence:
        print("Creating configuration matrices with hostname intelligence...")
    else:
        print("Creating configuration matrices...")
    print()
    
    all_matrices = {}
    all_analyses = {}
    total_config_entries = 0
    
    for server_type, hostname_files in server_hostname_files.items():
        matrix_df, hostname_analysis = process_server_type_matrix(
            server_type, hostname_files, args.output, enable_hostname_intelligence
        )
        all_matrices[server_type] = matrix_df
        all_analyses[server_type] = hostname_analysis
        total_config_entries += len(matrix_df)
    
    # Print overall summary
    print(f"\nOverall Processing Summary:")
    print(f"  Server types processed: {len(server_hostname_files)}")
    print(f"  Total hostnames processed: {total_hostnames}")
    print(f"  Total files found: {total_files}")
    print(f"  Total config entries: {total_config_entries}")
    
    # Print hostname intelligence summary across all server types
    if any(all_analyses.values()) and enable_hostname_intelligence:
        total_hostname_only = 0
        total_significant = 0
        total_same = 0
        
        for server_type, hostname_analysis in all_analyses.items():
            if hostname_analysis:
                hostname_only_count = sum(1 for v in hostname_analysis.values() if v == 'Hostname-Only')
                significant_count = sum(1 for v in hostname_analysis.values() if v == 'Significant')
                same_count = sum(1 for v in hostname_analysis.values() if v == 'Same')
                
                total_hostname_only += hostname_only_count
                total_significant += significant_count
                total_same += same_count
        
        print(f"\nOverall Hostname Intelligence Analysis:")
        print(f"  Same across hostnames: {total_same}")
        print(f"  Hostname-only differences: {total_hostname_only}")
        print(f"  Significant differences: {total_significant}")
        
        if total_hostname_only > 0:
            print(f"\nNote: {total_hostname_only} configuration entries differ only by hostname numbering.")
            print("These are highlighted in yellow in the Excel outputs.")
    
    print(f"\nGenerated separate diff matrix files for each server type:")
    base_name = os.path.splitext(args.output)[0]
    extension = os.path.splitext(args.output)[1] or '.xlsx'
    for server_type in server_hostname_files.keys():
        output_file = f"{base_name}_{server_type}_diff_matrix{extension}"
        print(f"  {output_file}")


if __name__ == "__main__":
    main()